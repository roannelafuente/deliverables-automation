import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import openpyxl
import csv
import os
import xlwings as xw
from datetime import datetime


class AutomatingDeliverables:
    def __init__(self, root):
        self.root = root
        self.root.title("Automating Deliverables")
        self.root.geometry("800x500")

        # Professional Neutral Theme
        self.bg_color = "#f5f5f5"
        self.fg_color = "#222222"
        self.entry_bg = "#ffffff"
        self.btn_bg = "#e0e0e0"
        self.btn_active = "#BEE395"

        self.root.configure(bg=self.bg_color)

        self.path_var = tk.StringVar()

        self.create_file_selection_frame()

        # Show filter selector immediately (empty at first)
        self.create_filter_selector([])

        # Then status box below it
        self.create_status_box()
        self.create_exit_button()


    def create_file_selection_frame(self):
        # File Selection frame with subtle border and spacing
        input_frame = tk.LabelFrame(
            self.root,
            text="File Selection",
            padx=10, pady=10,
            bd=2,
            relief="groove",
            font=("Segoe UI", 10, "bold")
        )
        input_frame.pack(fill="x", padx=15, pady=10)

        # Label inside the frame
        label = tk.Label(input_frame, text="Select CSV File:")
        label.pack(side="left", padx=(0, 10), pady=5)

        # Entry box inside the frame
        path_entry = tk.Entry(input_frame, textvariable=self.path_var,
                              bg="white", fg="black", insertbackground="black")
        path_entry.pack(side="left", padx=10, pady=5, fill="x", expand=True)

        # ✅ Convert button inside the same frame
        convert_btn = tk.Button(
            input_frame,
            text="Convert to Excel",
            width=18,
            command=self.convert_to_excel,
            bg=self.btn_bg,
            fg=self.fg_color,
            activebackground=self.btn_active
        )
        convert_btn.pack(side="right", padx=10, pady=5)
        # Browse button inside the frame
        browse_btn = tk.Button(input_frame, text="Browse", width=12, command=self.browse_file)
        browse_btn.pack(side="right", pady=5)


    def get_unique_c1_mark_values(raw_items):
        flat = []
        for item in raw_items:
            if isinstance(item, list):   # flatten nested lists
                flat.extend(item)
            elif item is not None:
                flat.append(item)

        # Strip whitespace but keep case
        cleaned = [str(i).strip() for i in flat if i]

        # Deduplicate while preserving order (case-sensitive)
        unique = list(dict.fromkeys(cleaned))
        return unique
    
    def create_filter_selector(self, items):
        # Pivot Filter Selection frame with subtle border and spacing
        filter_frame = tk.LabelFrame(
            self.root,
            text="Pivot Filter Selection",
            padx=10, pady=10,
            bd=2,
            relief="groove",
            font=("Segoe UI", 10, "bold")
        )

        filter_frame.pack(fill="x", padx=15, pady=10)

        tk.Label(filter_frame, text="Select C1_MARK:").pack(side="left", padx=5)

        # ✅ Clean and deduplicate items
        clean_items = [str(i) for i in items if i is not None]
        unique_items = list(dict.fromkeys(clean_items))

        self.filter_var = tk.StringVar()
        self.filter_dropdown = ttk.Combobox(
            filter_frame,
            textvariable=self.filter_var,
            values=unique_items,
            state="readonly",
            width=25
        )
        self.filter_dropdown.pack(side="left", padx=10)

        gen_pivot_btn = tk.Button(
            filter_frame,
            text="Generate Pivot Table",
            width=18,
            command=self.generate_pivot,
            bg=self.btn_bg,
            fg=self.fg_color,
            activebackground=self.btn_active
        )
        gen_pivot_btn.pack(side="left", padx=10)

        check_test_btn = tk.Button(
            filter_frame,
            text="Check End Test No",
            width=18,
            command=self.check_end_test,
            bg=self.btn_bg,
            fg=self.fg_color,
            activebackground=self.btn_active
        )
        check_test_btn.pack(side="left", padx=10)
            
    def create_status_box(self):
        status_frame = tk.LabelFrame(self.root, text="", padx=10, pady=10)
        status_frame.pack(fill="both", expand=True, padx=15, pady=10)

        self.status_box = tk.Text(
            status_frame,
            height=10,
            wrap="word",
            bg="white",
            fg="black",
            state="disabled"
        )
        self.status_box.pack(fill="both", expand=True)

    def create_exit_button(self):
        # Create an "invisible" frame with same background as root
        exit_frame = tk.Frame(self.root, bg=self.bg_color)
        exit_frame.pack(fill="x", side="bottom", padx=15, pady=5)

        # Place Exit button aligned right
        exit_btn = tk.Button(exit_frame, text="EXIT", width=12,
                             bg="#d32f2f", fg="white", command=self.root.destroy)
        exit_btn.pack(side="right", pady=10)

        clear_btn = tk.Button(exit_frame, text="Clear All", width=12,
                      command=self.clear_all,
                      bg=self.btn_bg, fg=self.fg_color, activebackground=self.btn_active)
        clear_btn.pack(side="right", padx=10)

    def show_status(self, message, color=None, clear=False):
        # Default to black unless explicitly set to red
        if color is None:
            color = "#000000"  # black

        self.status_box.config(state="normal")

        if clear:
            self.status_box.delete("1.0", "end")

        if message:
            self.status_box.insert("end", message + "\n")

            # Unique tag per line so colors don't overwrite
            line_tag = f"status_{self.status_box.index('end-2l')}"
            start_index = self.status_box.index("end-2l")
            end_index = self.status_box.index("end-1c")
            self.status_box.tag_add(line_tag, start_index, end_index)
            self.status_box.tag_config(line_tag, foreground=color)

        self.status_box.config(state="disabled")
        
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV files", "*.csv")]
        )
        if file_path:
            self.path_var.set(file_path)
            # Just show status that file is selected
            self.show_status(f"📂 Selected file:{file_path}", color="black")
            
    def convert_to_excel(self):
        file_path = self.path_var.get()
        if not file_path:
            self.show_status("⚠️ No file selected. Please browse for a CSV first.", color="#d32f2f")
            return

        try:
            # --- Convert CSV to Excel ---
            wb = openpyxl.Workbook()
            ws = wb.active

            sheet_name = os.path.splitext(os.path.basename(file_path))[0]
            ws.title = sheet_name[:31].replace(":", "_").replace("/", "_").replace("\\", "_")

            with open(file_path, newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                for r_idx, row in enumerate(reader, 1):
                    for c_idx, value in enumerate(row, 1):
                        try:
                            if value.isdigit():
                                ws.cell(row=r_idx, column=c_idx).value = int(value)
                            else:
                                ws.cell(row=r_idx, column=c_idx).value = float(value)
                        except ValueError:
                            ws.cell(row=r_idx, column=c_idx).value = value

            out_file = os.path.splitext(file_path)[0] + ".xlsx"
            wb.save(out_file)
            wb.close()

            # --- Open with xlwings to read filter items ---
            app = xw.App(visible=False)
            wb_xlw = app.books.open(out_file)
            sht = wb_xlw.sheets[0]

            # Find header row in Column G
            first_row = sht.range("G1").end("down").row
            header_value = sht.range(f"G{first_row}").value
            if header_value != "C1_MARK":
                self.show_status("❌ First non-empty cell in Column G is not 'C1_MARK'.", color="#d32f2f")
                wb_xlw.close()
                app.quit()
                return

            # Collect filter items from C1_MARK column
            last_row = sht.range((first_row, 7)).end("down").row
            raw_items = sht.range((first_row+1, 7), (last_row, 7)).value

            # ✅ Clean and deduplicate (case-sensitive)
            def get_unique_c1_mark_values(raw_items):
                flat = []
                for item in raw_items:
                    if isinstance(item, list):
                        flat.extend(item)
                    elif item is not None:
                        flat.append(item)
                cleaned = [str(i).strip() for i in flat if i]
                return list(dict.fromkeys(cleaned))  # preserves order, case-sensitive

            unique_items = get_unique_c1_mark_values(raw_items)

            # Update Combobox values
            self.filter_dropdown['values'] = unique_items

            # Keep workbook reference for later pivot creation
            self.out_file = out_file
            self.base_name = sheet_name
            wb_xlw.close()
            app.quit()

            self.show_status(
                f"✅ Conversion complete: CSV → .xlsx\nFile saved at: {out_file}\n\nFilter options loaded."
            )

        except Exception as e:
            self.show_status(f"❌ Error: {e}", color="#d32f2f")


    def generate_pivot(self):
        selected = self.filter_var.get()
        if not selected:
            self.show_status("⚠️ Please select a C1_MARK value first.", color="#d32f2f")
            return

        self.show_status(f"ℹ️ Generating pivot for selected C1_MARK: {selected}")

        app = None
        wb_xlw = None
        try:
            app = xw.App(visible=False)
            wb_xlw = app.books.open(self.out_file)
            sht = wb_xlw.sheets[self.base_name]

            # --- Find header row in Column G ---
            first_row = sht.range("G1").end("down").row
            header_value = sht.range(f"G{first_row}").value
            if str(header_value).strip().upper() != "C1_MARK":
                raise ValueError("First non-empty cell in Column G is not 'C1_MARK'")

            # --- Find ET header to the right of C1_MARK ---
            row_values = sht.range((first_row, 7), (first_row, sht.range((first_row, 7)).end("right").column)).value
            et_col = None
            for idx, val in enumerate(row_values, start=7):
                if str(val).strip().upper() == "ET":
                    et_col = idx
                    break
            if not et_col:
                raise ValueError("'ET' column not found to the right of C1_MARK")

            # --- Define pivot source range (C1_MARK → ET, down to last row in Column G) ---
            last_row = sht.range((first_row, 7)).end("down").row
            pivot_range = sht.range((first_row, 7), (last_row, et_col))

            # --- Create Pivot sheet ---
            pivot_sheet = wb_xlw.sheets.add("Pivot", after=sht)

            # --- Create pivot cache and table ---
            pivot_cache = wb_xlw.api.PivotCaches().Create(
                SourceType=1,  # xlDatabase
                SourceData=pivot_range.api
            )
            table_name = f"PivotTable_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            pivot_table = pivot_cache.CreatePivotTable(
                TableDestination=pivot_sheet.range("A3").api,
                TableName=table_name
            )

            # --- Filter: C1_MARK ---
            pf = pivot_table.PivotFields("C1_MARK")
            pf.Orientation = 3
            valid_items = [item.Name for item in pf.PivotItems()]
            if selected in valid_items:
                pf.CurrentPage = selected
                self.show_status(f"\nApplied filter: {selected}")
            else:
                self.show_status(f"⚠️ Selected '{selected}' not found in C1_MARK items {valid_items}", color="#d32f2f")
                return

            # --- Rows: ET ---
            pivot_table.PivotFields("ET").Orientation = 1

            # --- Values: Count of FT ---
            pivot_table.AddDataField(
                pivot_table.PivotFields("FT"),
                "Count of FT",
                -4112  # xlCount
            )

            # --- Fallout Table Logic ---
            data = pivot_sheet.range("A4").expand().value
            sheet = wb_xlw.sheets[self.base_name]
            theoretical_num = None
            for i, val in enumerate(sheet.range("A:A").value, start=1):
                if str(val).strip().upper() == "THEORETICAL_NUM":
                    theoretical_num = sheet.range((i, 1)).offset(0, 2).value
                    break

            fallout_table = []
            for row in data:
                if not row or not row[0] or str(row[0]).strip().lower() == "grand total":
                    continue
                et_val = str(int(row[0])) if isinstance(row[0], (int, float)) and float(row[0]).is_integer() else str(row[0])
                count_val = int(row[1]) if isinstance(row[1], (int, float)) and float(row[1]).is_integer() else row[1]
                fallout = (float(row[1]) / theoretical_num * 100) if theoretical_num else 0
                fallout_table.append([et_val, count_val, f"{fallout:.2f}%"])

            fallout_table.sort(key=lambda x: int(x[1]), reverse=True)
            grand_total_val = str(int(theoretical_num)) if isinstance(theoretical_num, (int, float)) and float(theoretical_num).is_integer() else str(theoretical_num)
            fallout_table.append(["Grand Total", grand_total_val, ""])


            # --- Write fallout table into Excel at D3 ---
            start_cell = pivot_sheet.range("D3")
            start_cell.value = ["End Test No.", "Count", "Fallout%"]
            pivot_sheet.range(start_cell.row+1, start_cell.column).value = fallout_table

            # --- Apply formatting ---
            last_row_ft = start_cell.row + len(fallout_table)  # last row of fallout table
            fallout_range = pivot_sheet.range(f"D3:F{last_row_ft}")

            # Header row (D3:F3)
            header_range = pivot_sheet.range("D3:F3")
            header_range.color = (192, 230, 245)   # #C0E6F5
            header_range.api.Font.Bold = True

            # First data row (D4:F4)
            first_data_range = pivot_sheet.range("D4:F4")
            first_data_range.color = (255, 159, 159)   # #FF9F9F
            first_data_range.api.Font.Bold = True

            # Grand Total row (last row)
            grand_total_range = pivot_sheet.range(f"D{last_row_ft}:F{last_row_ft}")
            grand_total_range.color = (192, 230, 245)   # #C0E6F5
            grand_total_range.api.Font.Bold = True

            # Borders: all inside thin, outside thick
            for i in range(7, 13):  # xlEdgeLeft=7, xlEdgeTop=8, xlEdgeBottom=9, xlEdgeRight=10, xlInsideVertical=11, xlInsideHorizontal=12
                fallout_range.api.Borders(i).Weight = 2  # xlThin inside
            for i in [7, 8, 9, 10]:
                fallout_range.api.Borders(i).Weight = 3  # xlThick outside

            # ✅ HOI: Center horizontally, middle vertically, reset indent
            fallout_range.api.HorizontalAlignment = -4108  # xlCenter
            fallout_range.api.VerticalAlignment = -4108    # xlCenter
            fallout_range.api.IndentLevel = 0              # no indentation
            
            wb_xlw.save()

            # --- Show fallout table in status box ---
            self.status_box.config(state="normal")
            #self.status_box.delete("1.0", tk.END)
            self.status_box.insert(tk.END, "\nPreview Table:\n")
            self.status_box.insert(tk.END, f"{'End Test No.':<15}{'Count':<10}{'Fallout%'}\n")
            self.status_box.insert(tk.END, "-" * 40 + "\n")
            for et_val, count_val, fallout_val in fallout_table:
                self.status_box.insert(
                    tk.END,
                    f"{str(et_val).strip():<15}{str(count_val):<10}{str(fallout_val)}\n"
                )
            self.status_box.config(state="disabled")

            self.show_status(f"\n✅ Pivot + Fallout table generated with filter {selected}")

        except Exception as e:
            self.show_status(f"❌ Error generating pivot/fallout: {e}", color="#d32f2f")

        finally:
            if wb_xlw:
                try: wb_xlw.close()
                except: pass
            if app:
                try: app.quit()
                except: pass


    def check_end_test(self):
        #self.show_status("\n🔍 Check End Test No clicked.")

        app = None
        wb_xlw = None
        try:
            app = xw.App(visible=False)
            wb_xlw = app.books.open(self.out_file)

            # --- Ensure Pivot sheet exists ---
            try:
                pivot_sheet = wb_xlw.sheets["Pivot"]
            except:
                pivot_sheet = wb_xlw.sheets.add("Pivot")

            data_sheet = wb_xlw.sheets[self.base_name]

            # --- Get highest fails End Test No from D4 (force string) ---
            raw_val = pivot_sheet.range("D4").value
            if raw_val is None:
                end_test_no = ""
            elif isinstance(raw_val, float) and raw_val.is_integer():
                end_test_no = str(int(raw_val))  # avoid .0
            else:
                end_test_no = str(raw_val).strip()

            self.show_status(f"\n🔍Checking End Test No.: {end_test_no}")

            # --- Find LOLIMIT row in Column F ---
            lolimit_row = data_sheet.range("F1").end("down").row
            lolimit_val = data_sheet.range(f"F{lolimit_row}").value
            if str(lolimit_val).strip().upper() != "LOLIMIT":
                raise ValueError("LOLIMIT not found in Column F")

            # --- Expand reference table ---
            ref_range = data_sheet.range((lolimit_row, 1)).expand("table")

            # --- Locate TESTNO column (Column B) ---
            testno_values = data_sheet.range(
                (lolimit_row + 1, 2),
                (lolimit_row + ref_range.rows.count - 1, 2)
            ).value

            # Normalize TESTNO values to strings
            testno_values = ["" if v is None else str(int(v)) if isinstance(v, float) and v.is_integer() else str(v).strip() for v in testno_values]

            found_row = None
            if end_test_no in testno_values:
                idx = testno_values.index(end_test_no) + lolimit_row + 1
                found_row = idx

            # --- Write End Test No. reference table into Excel at H3 ---
            start_cell = pivot_sheet.range("H3")
            start_cell.value = ["TSNO", "TESTNO", "COMMENT", "MODE", "HILIMIT", "LOLIMIT"]

            if found_row:
                row_values = data_sheet.range((found_row, 1), (found_row, 6)).value
                row_values = ["" if v is None else str(v).strip() for v in row_values]

                # Write row at H4:M4
                pivot_sheet.range(start_cell.row+1, start_cell.column).value = row_values

                # --- Apply formatting ---
                last_row = start_cell.row + 1
                ref_range_excel = pivot_sheet.range(f"H3:M{last_row}")

                # Header row (H3:M3)
                header_range = pivot_sheet.range("H3:M3")
                header_range.color = (192, 230, 245)   # #C0E6F5
                header_range.api.Font.Bold = True

                # Data row (H4:M4)
                data_range = pivot_sheet.range("H4:M4")
                data_range.color = (255, 255, 255)     # white background
                data_range.api.Font.Bold = False

                # Borders: all inside thin, outside thick
                for i in range(7, 13):  # xlEdgeLeft=7, xlEdgeTop=8, xlEdgeBottom=9, xlEdgeRight=10, xlInsideVertical=11, xlInsideHorizontal=12
                    ref_range_excel.api.Borders(i).Weight = 2
                for i in [7, 8, 9, 10]:
                    ref_range_excel.api.Borders(i).Weight = 3

                # Center horizontally and vertically
                ref_range_excel.api.HorizontalAlignment = -4108  # xlCenter
                ref_range_excel.api.VerticalAlignment = -4108    # xlCenter
                ref_range_excel.api.IndentLevel = 0

                wb_xlw.save()

                # --- Show End Test No. table in status box ---
                self.status_box.config(state="normal")
                self.status_box.insert(tk.END, "\nEnd Test No. Reference:\n")
                self.status_box.insert(
                    tk.END,
                    f"{'TSNO':<10}{'TESTNO':<10}{'COMMENT':<15}{'MODE':<10}{'HILIMIT':<10}{'LOLIMIT'}\n"
                )
                self.status_box.insert(tk.END, "-" * 70 + "\n")

                tsno, testno, comment, mode, hilimit, lolimit = row_values
                self.status_box.insert(
                    tk.END,
                    f"{tsno:<10}{testno:<10}{comment:<15}{mode:<10}{hilimit:<10}{lolimit}\n"
                )
                self.status_box.config(state="disabled")

                # --- Status message depending on limits ---
                if lolimit != "":
                    self.show_status("✅ Found with Limits")
                else:
                    # Amber for "no limit" warning
                    self.show_status("⚠️ Found with no Limit", color="#FFBF00")
            else:
                self.show_status("\n❌ No End Test No. found in the TESTNO Column", color="#d32f2f")

        except Exception as e:
            self.show_status(f"\n❌ Error checking End Test No: {e}", color="#d32f2f")

        finally:
            if wb_xlw:
                try:
                    wb_xlw.close()
                except:
                    pass
            if app:
                try:
                    app.quit()
                except:
                    pass
    def clear_all(self):
        # Reset file path
        self.path_var.set("")

        # Clear status box
        self.show_status("", clear=True)

        # Reset combobox selection and values
        if hasattr(self, "filter_dropdown"):
            self.filter_var.set("")                 # clear current selection
            self.filter_dropdown['values'] = []     # empty the dropdown list

# --- Run the App ---
if __name__ == "__main__":
    root = tk.Tk()
    app = AutomatingDeliverables(root)
    root.mainloop()
